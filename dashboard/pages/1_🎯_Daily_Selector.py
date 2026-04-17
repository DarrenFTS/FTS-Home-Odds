"""Page 1: Daily Bet Selector — Home Odds Portfolio"""
import streamlit as st
import pandas as pd
import os, sys
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from systems.all_systems import load_fixture_file, scan_all_systems, signals_to_dataframe, SYSTEM_DEFS
from dashboard.theme import SIDEBAR_CSS, sidebar_brand, G_PANEL, G_MID, G_ACCENT, G_TEST, G_BUF, G_LIVE

st.set_page_config(page_title="Daily Selector", page_icon="🎯", layout="wide")
sidebar_brand()

st.title("🎯 Daily Bet Selector")
st.markdown(
    f'<div style="background:{G_PANEL};border:1px solid {G_MID};border-left:3px solid {G_ACCENT};'
    f'border-radius:6px;padding:12px 16px;margin-bottom:16px;font-size:0.85rem;color:#b2dfdb">'
    f'Upload today\'s <strong style="color:#fff">FTS Advanced PreMatch Excel</strong> file. '
    f'All 7 systems will be scanned using home odds bands with a <strong style="color:{G_ACCENT}">±10% buffer</strong>. '
    f'<strong style="color:{G_LIVE}">🟢 LIVE</strong> = Lay U1.5 & Lay O3.5 (bet these). '
    f'<strong style="color:{G_TEST}">🔵 TEST</strong> = 5 other systems (track only).'
    f'</div>',
    unsafe_allow_html=True
)

uploaded = st.file_uploader("Upload FTS PreMatch file (Excel)", type=['xlsx','xls'])

if uploaded:
    with st.spinner("Running all 7 systems across home odds bands..."):
        try:
            tmp = f"/tmp/fixtures_{date.today()}.xlsx"
            with open(tmp, 'wb') as f:
                f.write(uploaded.read())
            fixtures = load_fixture_file(tmp)
            signals  = scan_all_systems(fixtures)
            df_all   = signals_to_dataframe(signals)
            fd       = fixtures['date'].dropna().iloc[0] if len(fixtures) else None
            date_str = fd.strftime('%A %d %B %Y') if fd is not None else str(date.today())
            date_file= fd.strftime('%Y%m%d') if fd is not None else str(date.today()).replace('-','')
        except Exception as e:
            st.error(f"Error loading file: {e}")
            st.stop()

    st.success(f"✅ {len(fixtures)} fixtures loaded — **{date_str}**")

    # ── KPI strip ──────────────────────────────────────────────────────────────
    live_sigs = [s for s in signals if s.live]
    test_sigs = [s for s in signals if not s.live]
    in_range  = [s for s in signals if not s.in_buffer]
    buffered  = [s for s in signals if s.in_buffer]

    from collections import Counter
    mc = Counter(s.system for s in signals)
    c1,c2,c3,c4,c5,c6,c7,c8 = st.columns(8)
    c1.metric("Total Selections", len(signals))
    c2.metric("🟢 LIVE",          len(live_sigs))
    c3.metric("🔵 TEST",          len(test_sigs))
    c4.metric("✅ In Range",       len(in_range))
    c5.metric("⚠️ Buffer Zone",   len(buffered))
    c6.metric("Lay U1.5",         mc.get("Lay U1.5", 0))
    c7.metric("Lay O3.5",         mc.get("Lay O3.5", 0))
    c8.metric("Other Systems",    sum(v for k,v in mc.items() if k not in ("Lay U1.5","Lay O3.5")))
    st.divider()

    if df_all.empty:
        st.warning("⚠️ No qualifying selections found in today's fixtures.")
        lgs = sorted(fixtures['comp'].dropna().unique()) if 'comp' in fixtures.columns else []
        if lgs:
            st.markdown(f"**Leagues in file:** {', '.join(lgs)}")
    else:
        # ── Filters ────────────────────────────────────────────────────────────
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            sys_filter = st.multiselect("System", sorted(df_all['System'].unique()),
                                         default=sorted(df_all['System'].unique()))
        with fc2:
            status_filter = st.multiselect("Status", ["🟢 LIVE","🔵 TEST"],
                                            default=["🟢 LIVE","🔵 TEST"])
        with fc3:
            buf_filter = st.multiselect("Range Status", ["✅ In Range","⚠️ Check KO"],
                                         default=["✅ In Range","⚠️ Check KO"])

        show = df_all[
            df_all['System'].isin(sys_filter) &
            df_all['Status'].isin(status_filter) &
            df_all['In Buffer'].isin(buf_filter)
        ].copy()

        # ── LIVE bets first, then TEST ─────────────────────────────────────────
        live_df = show[show['_live']].copy()
        test_df = show[~show['_live']].copy()

        def style_table(df):
            # Build display dataframe with exact column order
            rows = []
            for _, row in df.iterrows():
                date_val = row.get('Date', '')
                try:
                    date_val = pd.to_datetime(date_val).strftime('%d/%m/%Y')
                except Exception:
                    pass
                status_raw = str(row.get('In Buffer', ''))
                status_val = '\u26a0\ufe0f Check KO' if '\u26a0' in status_raw else '\u2705 In Range'
                rows.append({
                    'Date':                 date_val,
                    'Time':                 str(row.get('Time', '')),
                    'League (Competition)': str(row.get('League', '')),
                    'Home Team':            str(row.get('Home', '')),
                    'Away Team':            str(row.get('Away', '')),
                    'System':               str(row.get('System', '')),
                    'Bet':                  str(row.get('Bet', '')),
                    'Rule Range':           str(row.get('Rule Range', '')),
                    'Home Odds':            float(row.get('Home Odds', 0)),
                    'Lay Odds':             float(row.get('Lay Odds', 0)),
                    'Status':               status_val,
                })
            d = pd.DataFrame(rows)
            if d.empty:
                return d.style

            def color_system(v):
                s = str(v)
                if 'U1.5' in s or 'O3.5' in s:
                    return 'color:#2ecc71;font-weight:bold'
                return 'color:#5dade2'

            def color_status(v):
                s = str(v)
                if '\u2705' in s: return 'color:#2ecc71;font-weight:bold'
                if '\u26a0' in s: return 'color:#f39c12;font-weight:bold'
                return ''

            return (d.style
                    .format({'Home Odds': '{:.2f}', 'Lay Odds': '{:.2f}'})
                    .map(color_system, subset=pd.IndexSlice[:, ['System']])
                    .map(color_status, subset=pd.IndexSlice[:, ['Status']])
                    )

        if not live_df.empty:
            st.subheader("🟢 LIVE Bets — Place These")
            st.dataframe(style_table(live_df), use_container_width=True,
                         hide_index=True, height=min(600, 60+len(live_df)*35))

        if not test_df.empty:
            st.subheader("🔵 TEST Bets — Track Only (Do Not Place)")
            st.markdown(
                f'<div style="background:#0a2033;border:1px solid #1a4a6b;border-left:3px solid {G_TEST};'
                f'border-radius:6px;padding:10px 14px;margin-bottom:8px;color:#7fb3d3;font-size:0.82rem">'
                f'These selections qualify for TEST systems. Record results for performance tracking — '
                f'<strong style="color:#fff">do not place real stakes.</strong></div>',
                unsafe_allow_html=True
            )
            st.dataframe(style_table(test_df), use_container_width=True,
                         hide_index=True, height=min(600, 60+len(test_df)*35))

        st.divider()

        # ── Excel export ───────────────────────────────────────────────────────
        def build_excel(df_live, df_test, date_label):
            wb = Workbook()
            ws = wb.active
            ws.title = "LIVE Bets"

            def solid(h): return PatternFill("solid", fgColor=h)
            def bdr():
                s = Side(style="thin", color="CCCCCC")
                return Border(top=s, bottom=s, left=s, right=s)

            NAV="0D2B12"; GRN="2ECC71"; BLU="5DADE2"
            AMB="F39C12"; WHT="FFFFFF"; ALT="F0FFF4"
            BUFF="FFF9E6"; GRN_T="155A1E"

            def write_sheet(ws, df_in, is_live):
                colour = "2ECC71" if is_live else "5DADE2"
                ws.merge_cells("A1:K1")
                ws["A1"].value = f"FTS Home Odds — {'LIVE' if is_live else 'TEST'} Bets — {date_label}"
                ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
                ws["A1"].fill = solid(NAV)
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 26

                headers = ["Date","Time","League (Competition)","Home Team","Away Team","System","Bet",
                           "Rule Range","Home Odds","Lay Odds","Status"]
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(2, ci)
                    c.value = h
                    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                    c.fill = solid(colour if ci < 8 else NAV)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = bdr()
                ws.row_dimensions[2].height = 22

                # Column order: Date, Time, League, Home, Away, System, Bet, Rule Range, Home Odds, Lay Odds, Status
                df_cols = ['Date','Time','League','Home','Away','System','Bet',
                           'Rule Range','Home Odds','Lay Odds','In Buffer']
                df_use = df_in[[c for c in df_cols if c in df_in.columns]].reset_index(drop=True)

                for i, row in df_use.iterrows():
                    r = 3 + i
                    in_buf = '⚠️' in str(row.get('In Buffer', ''))
                    bg = BUFF if in_buf else (ALT if i%2==0 else WHT)
                    # Format date as DD/MM/YYYY
                    import pandas as _pd
                    date_val = row.get('Date','')
                    try:
                        date_val = _pd.to_datetime(date_val).strftime('%d/%m/%Y')
                    except: pass
                    vals = [date_val, row.get('Time',''), row.get('League',''),
                            row.get('Home',''), row.get('Away',''), row.get('System',''),
                            row.get('Bet',''), row.get('Rule Range',''),
                            row.get('Home Odds',''), row.get('Lay Odds',''),
                            '⚠️ Check KO' if in_buf else '✅ In Range']
                    for ci, v in enumerate(vals, 1):
                        cell = ws.cell(r, ci)
                        cell.value = v
                        cell.font = Font(name="Arial", size=10)
                        cell.fill = solid(bg)
                        # Left-align text columns: League(3), Home(4), Away(5), System(6), Rule Range(8)
                        cell.alignment = Alignment(horizontal="left" if ci in [3,4,5,6,8] else "center",
                                                   vertical="center")
                        cell.border = bdr()
                    # Colour status cell
                    sc = ws.cell(r, 11)
                    sc.font = Font(name="Arial", size=10, color=AMB if in_buf else GRN_T, bold=True)
                    ws.row_dimensions[r].height = 18

                widths = [13, 8, 26, 22, 22, 16, 8, 11, 11, 18, 20]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w
                ws.freeze_panes = "A3"

            write_sheet(ws, df_live, is_live=True)
            if not df_test.empty:
                ws2 = wb.create_sheet("TEST Bets")
                write_sheet(ws2, df_test, is_live=False)

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        xl_bytes = build_excel(live_df, test_df, date_str)

        dl_col, info_col = st.columns([1, 3])
        with dl_col:
            st.download_button(
                "⬇️ Download Excel",
                xl_bytes,
                file_name=f"FTS_HomeOdds_{date_file}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True
            )
        with info_col:
            st.info("Excel contains separate LIVE and TEST sheets with colour coding. "
                    "Green rows = in range · Amber rows = buffer zone (recheck at KO).")

    # ── Rules reference ────────────────────────────────────────────────────────
    with st.expander("📖 System Rules Reference — All 7 Systems"):
        for sys in SYSTEM_DEFS:
            badge = f"🟢 LIVE" if sys["live"] else "🔵 TEST"
            st.markdown(f"**{sys['name']}** {badge} — {sys['bet_type']} `{sys['bet_label']}` · Lay/Back odds: 1.00–5.99 · Buffer: ±10%")
            rules_md = " | ".join(f"{r[0]} ({r[1]}–{r[2]})" for r in sys["rules"])
            st.markdown(f"<small style='color:#81c784'>{rules_md}</small>", unsafe_allow_html=True)
            st.markdown("")
